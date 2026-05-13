import shutil
from datetime import date
import openpyxl
import pandas as pd
from config import (
    PC_SHEET_NAME,
    PC_HIS_NAME,
    PC_SHORT,
    PC_LAST_INV,
    FY_START_MONTH,
)
from core.logger import get_logger
from core.mrs_safety import atomic_replace, mrs_lock, backup_mrs, prune_backups

log = get_logger(__name__)


def get_financial_year(today: date) -> str:
    """Return FY string like '25-26'.

    month >= April → YY-(YY+1), else (YY-1)-YY
    """
    if today.month >= FY_START_MONTH:
        y1 = str(today.year)[-2:]
        y2 = str(today.year + 1)[-2:]
    else:
        y1 = str(today.year - 1)[-2:]
        y2 = str(today.year)[-2:]
    return f"{y1}-{y2}"


def get_next_invoice_no(party_row: pd.Series, today: date) -> str:
    """Format next invoice number as ShortCode-{seq:02d}/FY."""
    short_code = str(party_row[PC_SHORT]).strip()
    last_inv = party_row[PC_LAST_INV]

    try:
        last_seq = int(float(str(last_inv).strip()))
    except (ValueError, TypeError):
        last_seq = 0

    next_seq = last_seq + 1
    fy = get_financial_year(today)
    return f"{short_code}-{next_seq:02d}/{fy}"


def increment_invoice_no(his_name: str, master_path: str) -> None:
    """Atomically increment Last Invoice Number for a party in Party_Config.

    Wrapped in mrs_lock so two app instances cannot race on the same number.
    Wrapped in atomic_replace so a crash mid-save cannot leave a corrupt file.
    Backs up the live MRS first; backup is retained on failure for manual recovery.
    """
    log.info("increment_invoice_no: party=%s", his_name)
    backup_path = backup_mrs(master_path, tag=f"pre_inv_inc_{his_name[:20]}")

    with mrs_lock(master_path):
        with atomic_replace(master_path) as tmp_path:
            shutil.copy2(master_path, tmp_path)
            try:
                wb = openpyxl.load_workbook(tmp_path)
            except Exception as e:
                raise ValueError(f"Cannot open Master Rate Sheet to update invoice number: {e}")

            if PC_SHEET_NAME not in wb.sheetnames:
                raise ValueError(f"Sheet '{PC_SHEET_NAME}' not found in Master Rate Sheet.")

            ws = wb[PC_SHEET_NAME]

            header_row = None
            his_col = None
            inv_col = None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip() == PC_HIS_NAME:
                        header_row = cell.row
                        his_col = cell.column
                    if cell.value and str(cell.value).strip() == PC_LAST_INV:
                        inv_col = cell.column
                if header_row:
                    break

            if header_row is None or his_col is None or inv_col is None:
                raise ValueError(
                    f"Could not locate '{PC_HIS_NAME}' or '{PC_LAST_INV}' column in Party_Config sheet."
                )

            his_name_stripped = his_name.strip()
            target_row = None
            for row in ws.iter_rows(min_row=header_row + 1):
                cell_val = row[his_col - 1].value
                if cell_val and str(cell_val).strip() == his_name_stripped:
                    target_row = row
                    break

            if target_row is None:
                raise ValueError(
                    f"Party '{his_name}' not found in Party_Config sheet — cannot update invoice number."
                )

            inv_cell = target_row[inv_col - 1]
            try:
                current = int(float(str(inv_cell.value).strip()))
            except (TypeError, ValueError):
                current = 0
            inv_cell.value = current + 1
            log.info("Invoice number for %s: %d → %d", his_name, current, current + 1)

            try:
                wb.save(tmp_path)
            except Exception as e:
                raise ValueError(f"Could not save Master Rate Sheet after updating invoice number: {e}")
    prune_backups()
