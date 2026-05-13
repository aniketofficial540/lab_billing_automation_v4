import shutil
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import pandas as pd
from config import (
    WORKING_TEMPLATE_PATH,
    WS_ROW_PARTY_NAME,
    WS_COL_PARTY_NAME,
    WS_COL_BILLING_PERIOD,
    WS_ROW_DATA_START,
    WS_COL_SNO,
    WS_COL_COMPANY,
    WS_COL_PATIENT,
    WS_COL_MRD,
    WS_COL_BILL_NO,
    WS_COL_DATE,
    WS_COL_CPT,
    WS_COL_SERVICE,
    WS_COL_MRP,
    WS_COL_DISCOUNT,
    WS_COL_BILL_AMT,
    WS_FILL_WHITE,
    WS_FILL_BLUE,
    WS_FILL_GREEN1,
    WS_FILL_GREEN2,
    WS_FILL_UNMATCH,
    WS_FILL_TOTAL,
    WS_FILL_SEPARATOR,
    CSV_COL_DATE,
)


def _fill(argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=argb)


def _safe_set(ws, row: int, col: int, value=None, fill=None, font=None,
              number_format=None, alignment=None):
    """Write to a cell only if it is not a non-anchor MergedCell."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return  # non-anchor of a merged range — skip
    if value is not None:
        cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if number_format is not None:
        cell.number_format = number_format
    if alignment is not None:
        cell.alignment = alignment


def _unmerge_data_area(ws, first_row: int):
    """Remove any merged regions from the template that touch rows >= first_row."""
    to_remove = [
        str(mr) for mr in ws.merged_cells.ranges
        if mr.max_row >= first_row
    ]
    for ref in to_remove:
        ws.unmerge_cells(ref)


def _fmt_date(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
        return ""
    s = str(val).strip()
    from datetime import datetime
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return s


def write_working_sheet(
    matched_df: pd.DataFrame,
    unmatched_df: pd.DataFrame,
    party_full_name: str,
    billing_period: str,
    out_path: str,
) -> None:
    """Copy Working Sheet template and fill with data."""
    shutil.copy2(WORKING_TEMPLATE_PATH, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # Remove template merges that overlap the data area so we can write freely
    _unmerge_data_area(ws, WS_ROW_DATA_START)

    # Party name and billing period (row 3 — may be merged in template, write to anchors)
    _safe_set(ws, WS_ROW_PARTY_NAME, WS_COL_PARTY_NAME, value=party_full_name)
    _safe_set(ws, WS_ROW_PARTY_NAME, WS_COL_BILLING_PERIOD, value=billing_period)

    current_row = WS_ROW_DATA_START

    # Write matched rows
    for idx, (_, row) in enumerate(matched_df.iterrows()):
        is_even = idx % 2 == 1
        fill_ab = _fill(WS_FILL_BLUE if is_even else WS_FILL_WHITE)
        fill_ijk = _fill(WS_FILL_GREEN2 if is_even else WS_FILL_GREEN1)
        _write_row(ws, current_row, row, fill_ab, fill_ijk, is_unmatched=False)
        current_row += 1

    # Separator row
    sep_row = current_row
    _safe_set(
        ws, sep_row, 1,
        value="TESTS NOT IN AGREEMENT",
        fill=_fill(WS_FILL_SEPARATOR),
        font=Font(bold=True, color="FF000000"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    for col in range(2, 12):
        _safe_set(ws, sep_row, col, fill=_fill(WS_FILL_SEPARATOR))
    ws.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=11)
    current_row += 1

    # Write the bottom section.
    # Rows that have confirmed rates (filled by MRS lookup or by Op2) get the
    # SAME alternating white/blue + green colours as the matched section so
    # the sheet reads as one cohesive table — the "TESTS NOT IN AGREEMENT"
    # separator is what marks them as logically separate, not a different fill.
    # Only genuinely-blank rows (rate never confirmed) keep the peach UNMATCH
    # highlight as a visual flag for follow-up.
    for idx, (_, row) in enumerate(unmatched_df.iterrows()):
        bill_amt = row.get("Bill_Amount")
        has_rates = bill_amt is not None and not (isinstance(bill_amt, float) and pd.isna(bill_amt))
        is_even = idx % 2 == 1
        if has_rates:
            fill_ab  = _fill(WS_FILL_BLUE   if is_even else WS_FILL_WHITE)
            fill_ijk = _fill(WS_FILL_GREEN2 if is_even else WS_FILL_GREEN1)
        else:
            fill_ab  = _fill(WS_FILL_UNMATCH)
            fill_ijk = _fill(WS_FILL_UNMATCH)
        _write_row(ws, current_row, row, fill_ab, fill_ijk, is_unmatched=not has_rates)
        current_row += 1

    # TOTAL row — sum all data rows (matched + confirmed unmatched); separator has no K value so safe
    total_row = current_row
    for col in range(1, 12):
        _safe_set(ws, total_row, col, fill=_fill(WS_FILL_TOTAL))
    _safe_set(
        ws, total_row, 1,
        value="TOTAL",
        fill=_fill(WS_FILL_TOTAL),
        font=Font(bold=True),
        alignment=Alignment(horizontal="center"),
    )
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=10)

    k_col = get_column_letter(WS_COL_BILL_AMT)
    sum_formula = f"=SUM({k_col}{WS_ROW_DATA_START}:{k_col}{total_row - 1})"
    _safe_set(
        ws, total_row, WS_COL_BILL_AMT,
        value=sum_formula,
        fill=_fill(WS_FILL_TOTAL),
        font=Font(bold=True),
        number_format="#,##0.00",
    )

    try:
        wb.save(out_path)
    except Exception as e:
        raise ValueError(f"Could not save Working Sheet: {e}")


def _write_row(
    ws,
    row_num: int,
    row: pd.Series,
    fill_ab: PatternFill,
    fill_ijk: PatternFill,
    is_unmatched: bool,
) -> None:
    def _val(col_name):
        v = row.get(col_name)
        return "" if (v is None or (isinstance(v, float) and pd.isna(v))) else v

    col_vals = [
        (WS_COL_SNO,      _val("S.No.")),
        (WS_COL_COMPANY,  _val("Company Name")),
        (WS_COL_PATIENT,  _val("Patient Name")),
        (WS_COL_MRD,      _val("MRD Number")),
        (WS_COL_BILL_NO,  _val("Bill Number")),
        (WS_COL_DATE,     _fmt_date(_val(CSV_COL_DATE))),
        (WS_COL_CPT,      _val("cpt_code")),
        (WS_COL_SERVICE,  _val("Service Name")),
    ]
    for col, val in col_vals:
        _safe_set(ws, row_num, col, value=val, fill=fill_ab)

    if not is_unmatched:
        for col, raw in [
            (WS_COL_MRP,      row.get("MRP")),
            (WS_COL_DISCOUNT, row.get("Discount")),
            (WS_COL_BILL_AMT, row.get("Bill_Amount")),
        ]:
            if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
                _safe_set(ws, row_num, col, value=float(raw),
                          fill=fill_ijk, number_format="#,##0.00")
            else:
                _safe_set(ws, row_num, col, fill=fill_ijk)
    else:
        for col in (WS_COL_MRP, WS_COL_DISCOUNT, WS_COL_BILL_AMT):
            _safe_set(ws, row_num, col, fill=fill_ijk)


def calculate_total(matched_df: pd.DataFrame) -> float:
    if matched_df.empty or "Bill_Amount" not in matched_df.columns:
        return 0.0
    return float(matched_df["Bill_Amount"].dropna().astype(float).sum())
