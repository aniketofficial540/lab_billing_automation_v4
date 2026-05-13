import re
from datetime import date
import openpyxl
import pandas as pd
from config import (
    MR_SHEET_NAME,
    PC_SHEET_NAME,
    PC_HIS_NAME,
    PC_REMARKS,
    COL_TEST_CODE,
    COL_TEST_NAME,
    COL_MRP,
    MR_HEADER_ROW,
)
from core.logger import get_logger

log = get_logger(__name__)


def parse_addendum_excel(path: str) -> list[dict]:
    """Parse an Annexure-B format Excel file.

    Flexible column detection: looks for 'test code'/'code'/'cpt' and
    'discounted price'/'price'/'bill amount'/'bill_amount'.

    Handles files with a preamble (hospital header, party info rows) before
    the actual column header row by scanning to find it automatically.

    Returns list of dicts with keys: test_code, test_name, discounted_price
    """
    # First pass: read without a header assumption to locate the real header row.
    # Annexure-B files often have 4–6 preamble rows (hospital name, address,
    # party name, execution date, blank) before the actual column headers appear.
    try:
        raw = pd.read_excel(path, header=None, dtype=str)
    except Exception as e:
        raise ValueError(f"Cannot read addendum Excel file: {e}")

    header_row_idx = None
    for idx, row in raw.iterrows():
        row_vals = [str(v).strip().lower() for v in row if pd.notna(v) and str(v).strip()]
        has_code  = any("code" in v or "cpt" in v for v in row_vals)
        has_price = any("price" in v or "amount" in v or "discount" in v for v in row_vals)
        if has_code and has_price:
            header_row_idx = idx
            break

    if header_row_idx is None:
        raise ValueError(
            "Addendum Excel: could not locate the column header row. "
            "Expected a row containing 'Test Code' and 'Discounted Price' headings. "
            "Row values found in file: "
            + "; ".join(
                str(raw.iloc[i, 0]) for i in range(min(10, len(raw)))
            )
        )

    try:
        df = pd.read_excel(path, header=int(header_row_idx), dtype=str)
    except Exception as e:
        raise ValueError(f"Cannot read addendum Excel file: {e}")

    df.columns = [str(c).strip() for c in df.columns]
    col_lower = {c.lower(): c for c in df.columns}

    # Find test_code column
    tc_col = None
    for pattern in ["test code", "cpt code", "cpt", "code", "test_code"]:
        if pattern in col_lower:
            tc_col = col_lower[pattern]
            break
    if tc_col is None:
        # Try partial match
        for k, v in col_lower.items():
            if "code" in k or "cpt" in k:
                tc_col = v
                break
    if tc_col is None:
        raise ValueError(
            "Addendum Excel: could not find a 'Test Code' or 'CPT Code' column. "
            "Column headers found: " + ", ".join(df.columns.tolist())
        )

    # Find test_name column
    tn_col = None
    for pattern in ["test name", "service name", "test_name", "service", "name"]:
        if pattern in col_lower:
            tn_col = col_lower[pattern]
            break
    if tn_col is None:
        for k, v in col_lower.items():
            if "name" in k or "service" in k or "description" in k:
                tn_col = v
                break

    # Find discounted price column
    dp_col = None
    for pattern in [
        "discounted price",
        "discount price",
        "discounted_price",
        "bill amount",
        "bill_amount",
        "price",
        "amount",
    ]:
        if pattern in col_lower:
            dp_col = col_lower[pattern]
            break
    if dp_col is None:
        for k, v in col_lower.items():
            if "price" in k or "amount" in k or "bill" in k or "discount" in k:
                dp_col = v
                break
    if dp_col is None:
        raise ValueError(
            "Addendum Excel: could not find a 'Discounted Price' or 'Bill Amount' column. "
            "Column headers found: " + ", ".join(df.columns.tolist())
        )

    results = []
    for _, row in df.iterrows():
        tc = str(row[tc_col]).strip() if pd.notna(row[tc_col]) else ""
        if not tc or tc.lower() == "nan":
            continue
        tn = str(row[tn_col]).strip() if (tn_col and pd.notna(row[tn_col])) else ""
        dp_raw = str(row[dp_col]).strip() if pd.notna(row[dp_col]) else ""
        # Extract the first number from the cell — handles prefixes like <200, ₹200, ~200
        m = re.search(r'\d[\d,]*\.?\d*', dp_raw)
        if not m:
            raise ValueError(
                f"Addendum Excel: row for CPT '{tc}' has a non-numeric discounted price: '{dp_raw}'."
            )
        try:
            dp = float(m.group().replace(",", ""))
        except ValueError:
            raise ValueError(
                f"Addendum Excel: row for CPT '{tc}' has a non-numeric discounted price: '{dp_raw}'."
            )
        results.append({"test_code": tc, "test_name": tn, "discounted_price": dp})

    if not results:
        raise ValueError("Addendum Excel: no valid data rows found.")

    return results


def compute_addendum_preview(
    addendum_rows: list[dict],
    mr_column: str,
    master_path: str,
) -> list[dict]:
    """Read current MRS prices and return what the addendum will change — no writes.

    Returns list of dicts:
        test_code, test_name, old_price (None = new row), new_price, is_new_row
    """
    try:
        wb = openpyxl.load_workbook(master_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open Master Rate Sheet: {e}")

    if MR_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{MR_SHEET_NAME}' not found in Master Rate Sheet.")

    mrs = wb[MR_SHEET_NAME]
    excel_header_row = MR_HEADER_ROW + 1

    col_index_map: dict[str, int] = {}
    for cell in list(mrs.iter_rows(min_row=excel_header_row, max_row=excel_header_row))[0]:
        if cell.value:
            col_index_map[str(cell.value).strip()] = cell.column

    tc_col_idx = col_index_map.get(COL_TEST_CODE)
    party_col_idx = col_index_map.get(mr_column)
    mrp_col_idx = col_index_map.get(COL_MRP)

    if not tc_col_idx:
        raise ValueError(f"'{COL_TEST_CODE}' column not found in MRS header.")
    if not party_col_idx:
        raise ValueError(f"Party column '{mr_column}' not found in MRS headers.")

    # test_code → (current party price, current MRP)
    info_map: dict[str, tuple[float | None, float | None]] = {}
    for row in mrs.iter_rows(min_row=excel_header_row + 1):
        tc_val = row[tc_col_idx - 1].value
        if not tc_val:
            continue
        try:
            party_val = row[party_col_idx - 1].value
            party_f = float(party_val) if party_val is not None else None
        except (ValueError, TypeError):
            party_f = None
        try:
            mrp_val = row[mrp_col_idx - 1].value if mrp_col_idx else None
            mrp_f = float(mrp_val) if mrp_val is not None else None
        except (ValueError, TypeError):
            mrp_f = None
        info_map[str(tc_val).strip()] = (party_f, mrp_f)

    wb.close()

    changes = []
    for entry in addendum_rows:
        tc = str(entry["test_code"]).strip()
        tn = str(entry.get("test_name", ""))
        new_price = float(entry["discounted_price"])
        is_new = tc not in info_map
        old_price, mrp = info_map.get(tc, (None, None))
        changes.append({
            "test_code":  tc,
            "test_name":  tn,
            "old_price":  old_price if not is_new else None,
            "new_price":  new_price,
            "mrp":        mrp,           # current MRS MRP (None if unknown / new test)
            "is_new_row": is_new,
        })
    return changes


def validate_addendum_against_mrp(changes: list[dict]) -> list[str]:
    """Hard-reject rule (spec §4.2): discounted price must not exceed MRP.

    Skips tests where MRP is unknown (new tests, or MRP column blank).
    Returns list of human-readable error messages — empty list = OK.
    """
    errors: list[str] = []
    for ch in changes:
        if ch["is_new_row"]:
            continue  # no MRP yet — handled separately by employee
        mrp = ch.get("mrp")
        if mrp is None:
            continue  # MRP not stored in MRS for this test — cannot enforce
        if ch["new_price"] > mrp + 0.01:  # tolerance for floating-point noise
            errors.append(
                f"  • {ch['test_code']} ({ch['test_name'][:40]}): "
                f"discounted ₹{ch['new_price']:,.2f} exceeds MRP ₹{mrp:,.2f}"
            )
    if errors:
        log.warning(
            "Addendum validation rejected %d test(s) for MRP > Discounted Price",
            len(errors),
        )
    return errors


def apply_addendum(
    addendum_rows: list[dict],
    mr_column: str,
    his_name: str,
    effective_date_actual: date,
    master_path: str,
) -> dict:
    """Write addendum prices into MRS party column.

    Tests already in MRS → cell updated.
    Tests NOT in MRS → new row appended (Test Code, Test Name, party price).
    Remarks in Party_Config updated with actual effective date.
    All writes go through a temp file + os.replace (atomic on the same volume).
    Returns {"updated": N, "added": M}.
    """
    import shutil
    from core.mrs_safety import atomic_replace

    log.info(
        "apply_addendum: party=%s, column=%s, %d rows, effective=%s",
        his_name, mr_column, len(addendum_rows), effective_date_actual,
    )

    with atomic_replace(master_path) as tmp_path:
        # Start from a copy of the live MRS, then mutate the copy
        shutil.copy2(master_path, tmp_path)
        try:
            wb = openpyxl.load_workbook(tmp_path)
        except Exception as e:
            raise ValueError(f"Cannot open Master Rate Sheet: {e}")

        if MR_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Sheet '{MR_SHEET_NAME}' not found in Master Rate Sheet.")

        mrs = wb[MR_SHEET_NAME]
        excel_header_row = MR_HEADER_ROW + 1

        header_cells = list(mrs.iter_rows(min_row=excel_header_row, max_row=excel_header_row))[0]
        col_index_map: dict[str, int] = {}
        for cell in header_cells:
            if cell.value:
                col_index_map[str(cell.value).strip()] = cell.column

        if COL_TEST_CODE not in col_index_map:
            raise ValueError(f"'{COL_TEST_CODE}' column not found in '{MR_SHEET_NAME}' header row.")
        if mr_column not in col_index_map:
            raise ValueError(f"Party column '{mr_column}' not found in Master Rate Sheet headers.")

        tc_col_idx   = col_index_map[COL_TEST_CODE]
        tn_col_idx   = col_index_map.get(COL_TEST_NAME)
        party_col_idx = col_index_map[mr_column]

        row_map: dict[str, int] = {}
        for row in mrs.iter_rows(min_row=excel_header_row + 1):
            tc_cell = row[tc_col_idx - 1]
            if tc_cell.value:
                row_map[str(tc_cell.value).strip()] = tc_cell.row

        updated = 0
        added   = 0

        for entry in addendum_rows:
            tc = str(entry["test_code"]).strip()
            tn = str(entry.get("test_name", ""))
            dp = float(entry["discounted_price"])

            if tc in row_map:
                mrs.cell(row=row_map[tc], column=party_col_idx).value = dp
                updated += 1
            else:
                new_row = mrs.max_row + 1
                mrs.cell(row=new_row, column=tc_col_idx).value = tc
                if tn_col_idx:
                    mrs.cell(row=new_row, column=tn_col_idx).value = tn
                mrs.cell(row=new_row, column=party_col_idx).value = dp
                row_map[tc] = new_row  # prevent duplicates in same batch
                added += 1

        _update_remarks(wb, his_name, effective_date_actual)

        try:
            wb.save(tmp_path)
        except Exception as e:
            raise ValueError(f"Could not save Master Rate Sheet after applying addendum: {e}")

    log.info("apply_addendum done: updated=%d, added=%d", updated, added)
    return {"updated": updated, "added": added}


def _update_remarks(wb: openpyxl.Workbook, his_name: str, effective_date: date) -> None:
    """Write 'Addendum applied DD-Mon-YYYY' to Party_Config Remarks cell."""
    if PC_SHEET_NAME not in wb.sheetnames:
        return

    ws = wb[PC_SHEET_NAME]
    date_str = effective_date.strftime("%d-%b-%Y")
    remarks_text = f"Addendum applied {date_str}"

    # Locate header row
    header_row = None
    his_col = None
    rem_col = None

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == PC_HIS_NAME:
                header_row = cell.row
                his_col = cell.column
            if cell.value and str(cell.value).strip() == PC_REMARKS:
                rem_col = cell.column
        if header_row:
            break

    if not (header_row and his_col and rem_col):
        return

    his_stripped = his_name.strip()
    for row in ws.iter_rows(min_row=header_row + 1):
        val = row[his_col - 1].value
        if val and str(val).strip() == his_stripped:
            row[rem_col - 1].value = remarks_text
            break
