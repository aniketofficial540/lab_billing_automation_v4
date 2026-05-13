"""
core/master_rate_editor.py — In-app MRS editing and replacement.

All write paths go through atomic_replace + mrs_lock so the live MRS
is never left in a half-written state.
"""

import os
import shutil
import pandas as pd
import openpyxl
from config import (
    MASTER_RATE_PATH, MR_SHEET_NAME, PC_SHEET_NAME,
    MR_HEADER_ROW, COL_TEST_CODE, COL_TEST_NAME, COL_MRP,
)
from core.logger import get_logger
from core.mrs_safety import atomic_replace, mrs_lock, backup_mrs, prune_backups

log = get_logger(__name__)


def read_mrs_for_display() -> dict:
    """Read MRS for display/editing.
    Returns {headers: [str], rows: [[str,...]], party_columns: [str]}
    """
    df = pd.read_excel(
        MASTER_RATE_PATH, sheet_name=MR_SHEET_NAME,
        header=MR_HEADER_ROW, dtype=str,
    )
    df.columns = [str(c).strip() for c in df.columns]
    headers = list(df.columns)
    rows = [
        [str(v) if v is not None and str(v) != "nan" else "" for v in row]
        for row in df.values.tolist()
    ]
    non_party = {COL_TEST_CODE, COL_TEST_NAME, COL_MRP}
    party_columns = [c for c in headers if c not in non_party]
    return {"headers": headers, "rows": rows, "party_columns": party_columns}


def write_cells_batch(changes: dict) -> None:
    """Write a batch of cell edits to MRS atomically (under MRS lock).
    changes: {(row_idx, col_name): new_value}  (row_idx is 0-based data row)
    """
    log.info("write_cells_batch: %d cell change(s)", len(changes))
    if not changes:
        return

    backup_path = backup_mrs(MASTER_RATE_PATH, tag="pre_edit")
    try:
        with mrs_lock(MASTER_RATE_PATH):
            with atomic_replace(MASTER_RATE_PATH) as tmp_path:
                shutil.copy2(MASTER_RATE_PATH, tmp_path)
                wb = openpyxl.load_workbook(tmp_path)
                ws = wb[MR_SHEET_NAME]

                excel_header_row = MR_HEADER_ROW + 1
                header_row_cells = list(ws.iter_rows(
                    min_row=excel_header_row, max_row=excel_header_row, values_only=True
                ))[0]
                headers = [str(h).strip() if h is not None else "" for h in header_row_cells]
                col_index_map = {h: i + 1 for i, h in enumerate(headers) if h}
                data_start = excel_header_row + 1

                for (row_idx, col_name), new_value in changes.items():
                    if col_name not in col_index_map:
                        log.warning("write_cells_batch: unknown column %r — skipping", col_name)
                        continue
                    excel_row = data_start + row_idx
                    col_idx = col_index_map[col_name]
                    try:
                        ws.cell(row=excel_row, column=col_idx).value = (
                            float(new_value) if new_value not in ("", None) else None
                        )
                    except (ValueError, TypeError):
                        ws.cell(row=excel_row, column=col_idx).value = new_value

                try:
                    wb.save(tmp_path)
                except Exception as e:
                    raise ValueError(f"Could not save Master Rate Sheet: {e}")
        prune_backups()
    except Exception:
        # atomic_replace already discards the temp; nothing else to undo here
        log.error("write_cells_batch failed; backup retained at %s", backup_path)
        raise


def write_cell_edit(row_idx: int, col_name: str, new_value) -> None:
    """Write a single cell edit. Convenience wrapper around write_cells_batch."""
    write_cells_batch({(row_idx, col_name): new_value})


def replace_master_rate_sheet(new_file_path: str) -> None:
    """Validate structure of `new_file_path` and replace the live MRS atomically."""
    try:
        df = pd.read_excel(
            new_file_path, sheet_name=MR_SHEET_NAME,
            header=MR_HEADER_ROW, dtype=str,
        )
    except Exception as e:
        raise ValueError(f"Cannot read new file: {e}")
    df.columns = [str(c).strip() for c in df.columns]
    required = {COL_TEST_CODE, COL_TEST_NAME, COL_MRP}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"New Master Rate Sheet is missing required columns: {', '.join(sorted(missing))}."
        )
    if len(df) == 0:
        raise ValueError("New Master Rate Sheet has no data rows.")

    log.info("replace_master_rate_sheet: replacing with %s", new_file_path)
    # Backup the live file (if it exists) before swapping
    if os.path.exists(MASTER_RATE_PATH):
        backup_mrs(MASTER_RATE_PATH, tag="pre_replace")

    with mrs_lock(MASTER_RATE_PATH):
        with atomic_replace(MASTER_RATE_PATH) as tmp_path:
            shutil.copy2(new_file_path, tmp_path)
    prune_backups()
    log.info("replace_master_rate_sheet: done")
